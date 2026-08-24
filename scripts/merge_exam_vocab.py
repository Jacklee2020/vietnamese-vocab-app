# -*- coding: utf-8 -*-
"""清洗 A1-B2 考试词表并合并到词库_合并版.csv。

清洗步骤：
1. 修复字符间隙异常（如 t ì m c á ch -> tìm cách, t á o b ó n -> táo bón）
2. 修复粘连无空格词（如 chỗtrống -> chỗ trống, côgiáo -> cô giáo）
3. 剥离串入越南语字段的英文残留（如 nằm viện to be in -> nằm viện）
4. 从英文列提取错位的中文释义（如 a wife 妻子 -> 妻子）
5. 剔除无效 OCR 碎残字符与数字符号
6. 自动语义归类到 12 大标准化主题
7. 与现有词库查重，生成增补清单与合并全量词库
"""
import csv
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
APP_DIR = os.path.abspath(os.path.join(HERE, '..'))
ROOT_DIR = os.path.abspath(os.path.join(APP_DIR, '..'))

APP_CSV = os.path.join(APP_DIR, '词库_合并版.csv')
APP_XLSX = os.path.join(APP_DIR, '词库_合并版.xlsx')
OUT_NEW_CSV = os.path.join(APP_DIR, '新增清单_A1B2.csv')


def find_exam_csv():
    """解析待清洗源词表路径，支持命令行参数与多级目录回退。"""
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        return os.path.abspath(sys.argv[1])
    candidates = [
        os.path.join(ROOT_DIR, '越南语等级考试模拟-词表', 'A1-B2_新增词表_app未收录.csv'),
        os.path.join(APP_DIR, 'A1-B2_新增词表_app未收录.csv'),
        os.path.join(APP_DIR, 'scripts', 'A1-B2_新增词表_app未收录.csv'),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

# 12 大主题标准名称
C1  = '一、数字、量词与度量衡'
C2  = '二、时间、日期、月份与周期'
C3  = '三、人称代词、家庭亲属与职场社交'
C4  = '四、餐饮美食、水果特产与生活调料'
C5  = '五、交通出行、海关通关、地理与东南亚国家'
C6  = '六、购物百货、服装尺码、颜色与售后'
C7  = '七、酒店住宿、房屋租赁与日常家务'
C8  = '八、商务经贸、合同条款与物流生产'
C9  = '九、金融理财、股市交易与外汇结算'
C10 = '十、医疗健康、人体疾病与紧急求助'
C11 = '十一、核心动词、形容词、连词与高频成语'
C12 = '十二、常用口语短句（高频）'

CATEGORIES = [C1, C2, C3, C4, C5, C6, C7, C8, C9, C10, C11, C12]

# 常见无意义残片/单字符
JUNK_FRAGMENTS = {
    'nă', 'đ', 'Chù', 's', 'th', 'gi', 'phí', 'biết(đg.', 'quý,)',
    'Tripadvisor,)', 'Telegraph,(', 'chỉ*', '3', 'km', 'kg', 'm(mét)'
}

# 明确的 OCR 修正字典
SPECIFIC_CORRECTIONS = {
    't ì m c á ch': 'tìm cách',
    't á o b ó n': 'táo bón',
    'gi á m s á t': 'giám sát',
    'th í ch ... hơn': 'thích ... hơn',
    'tính đến chuyện gia đ ì nh': 'tính đến chuyện gia đình',
    'ph íđặt cọc/ tiền đặt cọc deposit': 'tiền đặt cọc',
    'chỗtrống': 'chỗ trống',
    'ởlại': 'ở lại',
    'lễhội': 'lễ hội',
    'đemđến': 'đem đến',
    'côgiáo': 'cô giáo',
    'TrungÁ': 'Trung Á',
    'bâygiờ': 'bây giờ',
    'chạyđến': 'chạy đến',
    'quanhđây': 'quanh đây',
    'yhệt': 'y hệt',
    'môtả': 'mô tả',
    'đầyđủ': 'đầy đủ',
    'gợiý': 'gợi ý',
    'hồ Hoàn Kiếm Ho': 'hồ Hoàn Kiếm',
    'ga (ga Hà Nội)': 'ga Hà Nội',
    'Bách Kho': 'đại học Bách Khoa',
    'ở đâu cũng (ở đâu ... cũng)': 'ở đâu cũng',
    'đi khám to examine': 'đi khám',
    'nằm viện to be in': 'nằm viện',
    'uống thuốc to': 'uống thuốc',
    'nói về to talk': 'nói về',
    'ý chính ma': 'ý chính',
    'là gì is': 'là gì',
    'không chịu được nhiệt cannot withst': 'không chịu được nhiệt',
    'cho ... vào to': 'cho ... vào',
    'hướng ... lên trên in': 'hướng ... lên trên',
    'xếp chung to': 'xếp chung',
    'cách đây lâu chưa how': 'cách đây lâu chưa',
    'so với to comp': 'so với',
    'bị ốm to be': 'bị ốm',
    'lúc đó at': 'lúc đó',
    'đi ngắm hoa nở to look at': 'đi ngắm hoa nở',
    'không ... gì not any': 'không ... gì',
    'không thua kém the': 'không thua kém',
    'Bộ Giáo dục và Đào tạo Ministry of Education': 'Bộ Giáo dục và Đào tạo',
    'trong điều kiện in': 'trong điều kiện',
    'Thành phố Hồ Chí Minh HoC': 'Thành phố Hồ Chí Minh',
    'bún miến hải sản sea': 'bún miến hải sản',
    'môn Giáo dục công dân Civic Educati': 'môn Giáo dục công dân',
    'những người con của họ their sons': 'những người con của họ',
    'đi biển to go to the b': 'đi biển',
    'anh em tôi my siblings': 'anh em tôi',
    'hãy cho tôi biết Please': 'hãy cho tôi biết',
    'nóng ơi là nóng it is': 'nóng ơi là nóng',
    'trung tâm chiếu phim quốc gi': 'trung tâm chiếu phim quốc gia',
    'ngày mai ( d.)': 'ngày mai',
    '5 (năm)': 'năm',
    'm(mét)': 'mét',
    'quý,)': 'quý',
    'chỉ*': 'chỉ',
    'biết(đg.': 'biết',
}

han_re = re.compile(r'[\u4e00-\u9fff]')


def clean_vietnamese(raw_vi: str) -> str:
    """正规化越南语词形。"""
    raw_vi = raw_vi.strip()
    if raw_vi in SPECIFIC_CORRECTIONS:
        return SPECIFIC_CORRECTIONS[raw_vi]

    vi = raw_vi
    # 修复间隙字母（如 d i ệ n -> diện）
    tokens = vi.split()
    if len(tokens) >= 3 and all(len(t) == 1 for t in tokens):
        vi = ''.join(tokens)
    elif ' ' in vi:
        vi = re.sub(
            r'(\b[a-zA-Zàáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệđìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵ]) (\b[a-zA-Zàáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệđìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵ])\b',
            r'\1\2', vi)
        vi = re.sub(
            r'(\b[a-zA-Zàáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệđìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵ]) (\b[a-zA-Zàáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệđìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵ])\b',
            r'\1\2', vi)

    # 去除尾部串入的常见英文词汇
    vi = re.sub(
        r'\s+(to\s+be\s+in|to\s+examine|to\s+talk|to\s+look\s+at|cannot\s+withst|to\s+comp|to\s+be|at|in|how|is|are|the|a|for|of|with|from|by|about|to|Please|sea|HoC|Civic\s+Educati|deposit|their\s+sons|not\s+any|same\s+as|drug|Ministry\s+of\s+Education|and\s+Training|the\s+condition|the\s+blooming\s+flowers|my\s+siblings)\b.*$',
        '', vi, flags=re.IGNORECASE).strip()

    # 清理行尾词性括号如 ( d.), (đg.), (t.)
    vi = re.sub(r'\s*\([a-z\s.]+\)$', '', vi).strip()
    return vi


def clean_chinese(zh: str, en: str, vi: str) -> str:
    """提取并正规化中文释义。"""
    zh = zh.strip()
    en = en.strip()

    # 若中文列为空但英文列包含中文（分列错位）
    if not zh and han_re.search(en):
        m = han_re.search(en)
        zh = en[m.start():].strip()

    # 针对缺失中文的特殊词条
    if not zh:
        if vi == 'thừa': zh = '多余；剩余'
        elif vi == 'trường': zh = '学校；场地'
        elif vi == 'cách đây lâu chưa': zh = '距今多久了？'
        elif vi == 'không ... gì': zh = '什么也不……'
        elif vi == 'đi biển': zh = '去海边'

    # 去除首尾异常标点
    zh = re.sub(r'^[)\s,;、]+', '', zh)
    zh = re.sub(r'[(\s,;、]+$', '', zh)
    zh = zh.replace('  ', ' ').strip()
    return zh


def classify_word(vi: str, zh: str, pos: str, lvl: str) -> str:
    """根据越南语词条和中文释义进行 12 大主题归类。"""
    # 常用口语短句（高频）
    if any(q in zh for q in ['？', '?', '！', '!']) or any(k in zh for k in ['吗？', '呢？', '吧', '请问', '走吧', '好吗', '没关系', '不用谢', '太好了', '怎么了', '怎么样']):
        return C12
    if len(vi.split()) >= 4 and any(k in zh for k in ['我', '你', '他', '她', '请', '别', '不要', '如果', '为了', '因为', '所以']):
        return C12

    # 数字、量词与度量衡
    if any(k in zh for k in ['公里', '公斤', '米', '厘米', '毫米', '升', '毫升', '平方', '公顷', '度', '克', '千克', '吨', '量词', '倍', '百分之', '第', '双', '打', '半', '几', '多', '少', '面积', '容量', '长短', '高度', '深度', '宽度']):
        if not any(k in zh for k in ['时间', '年', '月', '日', '天', '星期']):
            return C1

    # 时间、日期、月份与周期
    if any(k in zh for k in ['年', '月', '日', '天', '周', '星期', '礼拜', '时', '分', '秒', '早', '晚', '午', '夜', '晨', '暮', '春', '夏', '秋', '冬', '季', '世纪', '年代', '假期', '放假', '周末', '平时', '准时', '按时', '迟到', '提前', '推迟', '延期', '节日', '春节', '元旦', '时代', '时期', '阶段', '期间', '当时', '那时', '现在', '将来', '未来', '过去', '曾经', '刚才', '刚刚', '马上', '立刻', '顿时', '随时', '届时', '岁', '年纪']):
        return C2

    # 人称代词、家庭亲属与职场社交
    if any(k in zh for k in ['我', '你', '他', '她', '它', '我们', '你们', '他们', '爸', '妈', '父', '母', '哥', '姐', '弟', '妹', '爷', '奶', '叔', '伯', '姑', '舅', '姨', '夫', '妻', '儿', '女', '孙', '亲', '友', '朋', '同', '师', '生', '长', '老', '少', '男', '女', '人', '辈', '亲戚', '家庭', '家人', '亲属', '朋友', '同事', '同学', '老师', '学生', '领导', '老板', '经理', '员工', '同事', '夫妻', '夫妇', '邻居', '客气', '礼貌', '拜访', '打扰', '介绍', '认识', '打招呼', '告别', '约会', '聚会']):
        return C3

    # 餐饮美食、水果特产与生活调料
    if any(k in zh for k in ['吃', '喝', '饭', '菜', '汤', '面', '粉', '粥', '肉', '鱼', '虾', '蟹', '蛋', '奶', '茶', '酒', '水', '果', '蔬', '粮', '油', '盐', '酱', '醋', '糖', '辣', '甜', '酸', '苦', '咸', '香', '鲜', '熟', '生', '烤', '炸', '煎', '炒', '煮', '蒸', '炖', '焖', '拌', '餐', '饮', '宴', '席', '馆', '厅', '味', '春卷', '米线', '河粉', '面包', '蛋糕', '饼', '点心', '零食', '饮料', '咖啡', '果汁', '啤酒', '水果', '西瓜', '苹果', '香蕉', '橘子', '蔬菜', '调料', '鱼露', '花生', '腰果']):
        return C4

    # 交通出行、海关通关、地理与东南亚国家
    if any(k in zh for k in ['车', '船', '飞', '机', '站', '港', '路', '街', '道', '桥', '河', '江', '海', '湖', '山', '岛', '国', '省', '市', '县', '镇', '村', '区', '地', '方', '城', '关', '境', '票', '证', '照', '签', '航', '线', '程', '驾', '骑', '走', '行', '游', '观', '景', '导', '旅', '客', '乘', '搭', '转', '换', '到', '达', '离', '去', '回', '往', '返', '发', '通', '塞', '堵', '挤', '快', '慢', '速', '距', '河内', '胡志明', '岘港', '海防', '大叻', '芽庄', '下龙湾', '越南', '老挝', '柬埔寨', '泰国', '中国', '日本', '亚洲', '海关', '签证', '护照', '机票', '车站', '码头', '机场', '路线', '地图', '方向', '东', '南', '西', '北']):
        return C5

    # 购物百货、服装尺码、颜色与售后
    if any(k in zh for k in ['买', '卖', '购', '物', '店', '超', '市', '商', '场', '货', '品', '衣', '服', '裤', '裙', '鞋', '帽', '袜', '包', '袋', '尺', '码', '大', '中', '小', '号', '颜', '色', '红', '黄', '蓝', '绿', '白', '黑', '紫', '灰', '金', '银', '试', '穿', '换', '退', '折', '扣', '优', '惠', '便', '宜', '贵', '价', '格', '发', '票', '收', '据', '保', '修', '售', '后', '专', '柜', '牌', '商', '标', '质', '量']):
        return C6

    # 酒店住宿、房屋租赁与日常家务
    if any(k in zh for k in ['住', '宿', '房', '屋', '店', '馆', '租', '赁', '借', '还', '床', '被', '枕', '单', '门', '窗', '锁', '匙', '洗', '扫', '擦', '理', '洁', '净', '脏', '乱', '水', '电', '气', '暖', '调', '扇', '炉', '锅', '碗', '盆', '桶', '巾', '刷', '膏', '液', '皂', '洗', '衣', '卫', '生', '厨', '卧', '客', '厅', '阳', '台', '楼', '层', '梯', '院', '园', '舍', '寓', '宅', '家', '居', '具', '电', '器', '家', '务', '打', '扫', '收', '拾', '整', '理']):
        return C7

    # 商务经贸、合同条款与物流生产
    if any(k in zh for k in ['商', '贸', '企', '业', '公', '司', '合', '同', '约', '协', '谈', '判', '合', '作', '项', '目', '投', '资', '资', '金', '产', '品', '业', '务', '市', '场', '销', '售', '订', '单', '交', '货', '运', '输', '物', '流', '仓', '储', '进', '口', '出', '口', '关', '税', '配', '额', '招', '标', '投', '标', '中', '标', '协', '议', '章', '程', '规', '则', '制', '度', '管', '理', '经', '营', '生', '产', '制', '造', '加', '工', '工', '厂', '车', '间', '工', '人', '技', '术', '设', '备', '原', '料', '标', '准', '检', '验', '合', '格', '证', '书', '许', '可', '牌', '照', '经', '纪', '代', '理', '批', '发', '零', '售', '展', '览', '博', '览', '研', '讨', '峰', '会', '考', '察', '会', '见', '接', '待', '备', '忘', '录']):
        return C8

    # 金融理财、股市交易与外汇结算
    if any(k in zh for k in ['钱', '币', '金', '银', '铜', '钞', '券', '汇', '款', '账', '户', '存', '取', '贷', '还', '借', '息', '利', '率', '税', '费', '价', '值', '本', '利', '盈', '亏', '损', '益', '股', '票', '证', '券', '基', '金', '期', '货', '保', '险', '信', '用', '卡', '支', '票', '转', '账', '结', '算', '支', '付', '兑', '换', '汇', '率', '贬', '值', '升', '值', '通', '胀', '紧', '缩', '行', '情', '涨', '跌', '牛', '市', '熊', '市', '开', '户', '平', '仓', '爆', '仓', '理', '财', '分', '红', '股', '息', '资', '产', '负', '债', '净', '值', '财', '政', '金', '融', '银', '行', '保', '险']):
        return C9

    # 医疗健康、人体疾病与紧急求助
    if any(k in zh for k in ['医', '药', '病', '痛', '伤', '疾', '症', '院', '诊', '护', '针', '刀', '血', '骨', '皮', '心', '肝', '脾', '肺', '肾', '胃', '肠', '脑', '头', '脸', '眼', '耳', '鼻', '口', '牙', '舌', '喉', '胸', '背', '腰', '腹', '手', '脚', '腿', '健康', '感冒', '发烧', '咳嗽', '便秘', '腹泻', '呕吐', '失眠', '过敏', '发炎', '中毒', '流血', '打针', '吃药', '处方', '住院', '出院', '急救', '救护', '癌症', '血压', '血糖', '体温', '营养', '维生素', '疫苗', '病毒', '细菌', '隔离', '消毒', '病房', '挂号', '急诊']):
        return C10

    # 默认为核心词汇
    return C11


def main():
    # 1. 加载现有词库
    if not os.path.exists(APP_CSV):
        print(f"❌ 未找到现有词库：{APP_CSV}")
        return 1

    with open(APP_CSV, encoding='utf-8-sig') as f:
        existing_rows = list(csv.reader(f))[1:]

    existing_vi_map = {r[2].strip().lower(): r for r in existing_rows if len(r) >= 4}
    print(f"📖 现有词库读取完毕：{len(existing_rows)} 词")

    # 2. 读取并清洗考试未收录词表
    src_exam_csv = find_exam_csv()
    if not src_exam_csv or not os.path.exists(src_exam_csv):
        print("❌ 未找到考试源词表文件（A1-B2_新增词表_app未收录.csv）。")
        print("   用法：python3 scripts/merge_exam_vocab.py <源词表路径>")
        print("   或者将源文件放置在「越南语等级考试模拟-词表/」或仓库根目录。")
        return 1

    print(f"📄 读取待清洗考试源词表：{src_exam_csv}")
    with open(src_exam_csv, encoding='utf-8-sig') as f:
        raw_rows = list(csv.reader(f))[1:]

    cleaned_new_entries = []
    seen_vi_new = set()
    skipped_count = 0

    for idx, r in enumerate(raw_rows, 1):
        if not r or not any(r):
            continue
        raw_vi = r[0].strip()
        pos = r[1].strip() if len(r) > 1 else ''
        raw_zh = r[2].strip() if len(r) > 2 else ''
        raw_en = r[3].strip() if len(r) > 3 else ''
        lvl = r[4].strip() if len(r) > 4 else ''

        if raw_vi in JUNK_FRAGMENTS:
            skipped_count += 1
            continue

        vi = clean_vietnamese(raw_vi)
        if not vi or vi in JUNK_FRAGMENTS or len(vi) <= 1:
            skipped_count += 1
            continue

        zh = clean_chinese(raw_zh, raw_en, vi)
        if not zh:
            skipped_count += 1
            continue

        vi_key = vi.lower()
        if vi_key in existing_vi_map:
            skipped_count += 1
            continue

        if vi_key in seen_vi_new:
            skipped_count += 1
            continue

        seen_vi_new.add(vi_key)
        category = classify_word(vi, zh, pos, lvl)

        cleaned_new_entries.append({
            'category': category,
            'vi': vi,
            'zh': zh,
            'level': lvl,
            'pos': pos,
        })

    print(f"✨ 考试词表清洗完成：有效新增 {len(cleaned_new_entries)} 词（跳过/去重 {skipped_count} 条）")

    # 3. 写入新增清单 CSV（若有新增）
    if cleaned_new_entries:
        with open(OUT_NEW_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['序号', '分类', '越南语', '中文', '级别', '词性'])
            for i, item in enumerate(cleaned_new_entries, 1):
                writer.writerow([i, item['category'], item['vi'], item['zh'], item['level'], item['pos']])
        print(f"📄 新增清单已导出：{OUT_NEW_CSV}")
    else:
        print("ℹ️ 无新词需要增补（现有词库已包含全部条目）。")

    # 4. 构建合并后的全量词库
    merged_rows = []
    # 保持原顺序并重排序号
    seq = 1
    # 按照 12 大主题分组排序，保持词库整洁有序
    all_by_cat = {c: [] for c in CATEGORIES}

    # 加入现有词条
    for r in existing_rows:
        cat, vi, zh = r[1].strip(), r[2].strip(), r[3].strip()
        if cat in all_by_cat:
            all_by_cat[cat].append((vi, zh))
        else:
            all_by_cat[C11].append((vi, zh))

    # 加入新增词条
    for item in cleaned_new_entries:
        cat = item['category']
        all_by_cat[cat].append((item['vi'], item['zh']))

    for cat in CATEGORIES:
        for vi, zh in all_by_cat[cat]:
            merged_rows.append([seq, cat, vi, zh])
            seq += 1

    # 5. 写入更新后的 词库_合并版.csv
    with open(APP_CSV, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['序号', '分类', '越南语', '中文'])
        for r in merged_rows:
            writer.writerow(r)

    print(f"✅ 全量词库更新完毕：{len(merged_rows)} 词 · 12 大类")
    print(f"   目标文件：{APP_CSV}")

    # 6. 生成 XLSX (如果 openpyxl 可用)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "词库合并版"
        ws.append(["序号", "分类", "越南语", "中文"])
        hfill = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        for c in range(1, 5):
            cell = ws.cell(row=1, column=c)
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Border(*[Side(style="thin", color="BFBFBF")]*4)
        palette = ["FFFFFF", "EAF1F8", "F5F0E6"]
        ti = 0; prev_theme = None
        for i, (s, cat, vi, zh) in enumerate(merged_rows, 2):
            if cat != prev_theme:
                ti += 1; prev_theme = cat
            fill = PatternFill("solid", fgColor=palette[ti % len(palette)])
            for c, v in enumerate([s, cat, vi, zh], 1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.border = thin; cell.fill = fill
                cell.alignment = Alignment(vertical="center")
        for col, wd in zip("ABCD", [8, 38, 30, 28]):
            ws.column_dimensions[col].width = wd
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:D{len(merged_rows)+1}"
        wb.save(APP_XLSX)
        print(f"📊 Excel 版本已更新：{APP_XLSX}")
    except Exception as e:
        print(f"⚠️ 生成 Excel 出现异常（不影响 CSV）：{e}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
